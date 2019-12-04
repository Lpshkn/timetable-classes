import openpyxl as xl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment

COL_TITLE = 1
COL_TEACH = 2
COL_GROUP = 3
COL_CLASSROOM = 4

THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


def write(days):
    wb = xl.Workbook()
    sheet = wb.active

    max_len_title = max_len_teach = max_len_grp = max_len_cls = 0

    cur_row = 1
    for day in days:
        # Merge cells to setup the date
        sheet.merge_cells(start_column=COL_TITLE, end_column=COL_CLASSROOM, start_row=cur_row, end_row=cur_row)

        # Setup the date of the day at the top of each table block and increment st_row number
        cell = sheet.cell(cur_row, COL_TITLE)
        cell.value = day.date
        cell.alignment = Alignment(horizontal='center')

        # Setup title's borders
        for col in range(COL_TITLE, COL_CLASSROOM + 1):
            sheet.cell(cur_row, col).border = THIN_BORDER

        cur_row += 1

        for note in day.notes:
            # Setup end_row value to then merge cells
            e_row = cur_row + len(note.teachers) - 1
            sheet.merge_cells(start_row=cur_row, end_row=e_row, start_column=COL_TITLE, end_column=COL_TITLE)

            # Setup the title
            cell = sheet.cell(cur_row, COL_TITLE)
            cell.value = note.title
            cell.border = THIN_BORDER
            max_len_title = max(max_len_title, len(cell.value))

            # Setup each teacher with attributes into a row
            for teacher in note.teachers:
                sheet.cell(cur_row, COL_TITLE).border = THIN_BORDER

                cell = sheet.cell(cur_row, COL_TEACH)
                cell.value = teacher.name
                cell.border = THIN_BORDER
                max_len_teach = max(max_len_teach, len(cell.value))

                cell = sheet.cell(cur_row, COL_GROUP)
                cell.value = ', '.join(teacher.groups)
                cell.border = THIN_BORDER
                max_len_grp = max(max_len_grp, len(cell.value))

                cell = sheet.cell(cur_row, COL_CLASSROOM)
                cell.value = str(teacher.classroom)
                cell.border = THIN_BORDER
                max_len_cls = max(max_len_cls, len(cell.value))
                cur_row += 1

    # Setup the width of columns
    sheet.column_dimensions['A'].width = max_len_title + 2
    sheet.column_dimensions['B'].width = max_len_teach + 2
    sheet.column_dimensions['C'].width = max_len_grp + 2
    sheet.column_dimensions['D'].width = max_len_cls + 2

    wb.save('1.xlsx')
